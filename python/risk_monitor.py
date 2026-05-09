"""
risk_monitor.py — Feature 18: Risk-Based Monitoring
Scores each site (0–100) based on data quality, SAEs, and query age.
Save in: Mini_EDC_Project/python/risk_monitor.py

Run: python risk_monitor.py
"""

import os
import json
from datetime import datetime
from db_connection import get_conn, is_postgres

REPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "reports")
os.makedirs(REPORT_DIR, exist_ok=True)


# ── Scoring weights ───────────────────────────────────────────
# Total max = 100 points of risk (higher = worse)
WEIGHTS = {
    "critical_queries": 25,  # per critical open query
    "sae_pending": 30,  # per pending SAE report
    "major_queries": 10,  # per major open query
    "stale_queries_7d": 15,  # per query open 7+ days
    "unanswered_rate": 20,  # % of queries never answered × 0.2
}
MAX_SCORE = 100


def get_db():
    return get_conn()


def _fetchone(cursor_result):
    """Fetch a single scalar value from a fetchone() result."""
    row = cursor_result.fetchone()
    if row is None:
        return 0
    # psycopg2 returns tuples; sqlite3 with Row factory returns Row objects
    return row[0]


def get_sites():
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT siteid FROM subjects ORDER BY siteid"
    ).fetchall()
    conn.close()
    return [r[0] for r in rows]


def score_site(conn, siteid):
    """Compute risk score for a single site. Returns score dict."""
    ph = "%s" if is_postgres() else "?"
    score = 0
    factors = []

    # Stale query age expression — differs between Postgres and SQLite
    if is_postgres():
        stale_expr = "EXTRACT(EPOCH FROM (NOW() - created_at::timestamp)) / 86400 > 7"
    else:
        stale_expr = "julianday('now') - julianday(created_at) > 7"

    # 1. Critical open queries
    crit = _fetchone(conn.execute(
        f"SELECT COUNT(*) FROM queries WHERE siteid={ph} AND severity='Critical' AND status='Open'",
        (siteid,),
    ))
    pts = min(crit * WEIGHTS["critical_queries"], 40)
    score += pts
    if crit:
        factors.append(
            {
                "factor": "Critical open queries",
                "count": crit,
                "points": pts,
                "level": "HIGH",
            }
        )

    # 2. SAE pending report
    sae = _fetchone(conn.execute(
        f"""
        SELECT COUNT(*) FROM adverse_events ae
        JOIN subjects s USING(usubjid)
        WHERE s.siteid={ph} AND ae.aeser='Y' AND ae.report_flag='PENDING'
    """,
        (siteid,),
    ))
    pts = min(sae * WEIGHTS["sae_pending"], 40)
    score += pts
    if sae:
        factors.append(
            {
                "factor": "SAEs pending 24hr report",
                "count": sae,
                "points": pts,
                "level": "CRITICAL",
            }
        )

    # 3. Major open queries
    major = _fetchone(conn.execute(
        f"SELECT COUNT(*) FROM queries WHERE siteid={ph} AND severity='Major' AND status='Open'",
        (siteid,),
    ))
    pts = min(major * WEIGHTS["major_queries"], 20)
    score += pts
    if major:
        factors.append(
            {
                "factor": "Major open queries",
                "count": major,
                "points": pts,
                "level": "MEDIUM",
            }
        )

    # 4. Stale queries (7+ days)
    stale = _fetchone(conn.execute(
        f"""
        SELECT COUNT(*) FROM queries
        WHERE siteid={ph} AND status='Open'
          AND {stale_expr}
    """,
        (siteid,),
    ))
    pts = min(stale * WEIGHTS["stale_queries_7d"], 20)
    score += pts
    if stale:
        factors.append(
            {
                "factor": "Stale queries (7+ days)",
                "count": stale,
                "points": pts,
                "level": "MEDIUM",
            }
        )

    # 5. Unanswered rate
    total_q = _fetchone(conn.execute(
        f"SELECT COUNT(*) FROM queries WHERE siteid={ph}", (siteid,)
    ))
    open_q = _fetchone(conn.execute(
        f"SELECT COUNT(*) FROM queries WHERE siteid={ph} AND status='Open'", (siteid,)
    ))
    unans_rate = (open_q / total_q * 100) if total_q else 0
    pts = min(int(unans_rate * WEIGHTS["unanswered_rate"] / 100), 20)
    score += pts
    if pts:
        factors.append(
            {
                "factor": f"Unanswered rate ({unans_rate:.0f}%)",
                "count": open_q,
                "points": pts,
                "level": "LOW",
            }
        )

    # Cap at 100
    score = min(score, MAX_SCORE)

    # Risk category
    if score >= 60:
        category = "HIGH RISK"
        action = "Immediate on-site visit required"
        icon = "🔴"
    elif score >= 30:
        category = "MEDIUM RISK"
        action = "Remote monitoring review within 2 weeks"
        icon = "🟡"
    else:
        category = "LOW RISK"
        action = "Routine monitoring schedule"
        icon = "🟢"

    # Site stats
    subjects = _fetchone(conn.execute(
        f"SELECT COUNT(*) FROM subjects WHERE siteid={ph}", (siteid,)
    ))

    return {
        "siteid": siteid,
        "score": score,
        "category": category,
        "action": action,
        "icon": icon,
        "subjects": subjects,
        "critical_queries": crit,
        "major_queries": major,
        "sae_pending": sae,
        "stale_queries": stale,
        "unanswered_rate": round(unans_rate, 1),
        "factors": factors,
    }


def run_risk_assessment():
    print("\n" + "=" * 65)
    print("  FEATURE 18 — Risk-Based Monitoring Site Assessment")
    print("=" * 65)

    conn = get_db()
    if not conn:
        return []

    sites = get_sites()
    if not sites:
        print("[RISK] No sites found in database.")
        return []

    results = []
    for siteid in sites:
        results.append(score_site(conn, siteid))
    conn.close()

    # Sort by score descending
    results.sort(key=lambda x: x["score"], reverse=True)

    # ── Console report ────────────────────────────────────────
    print(
        f"\n{'Site':<10} {'Score':>6} {'Category':<15} {'Subj':>5} {'Crit':>5} {'SAE':>5} {'Stale':>6}  Action"
    )
    print("─" * 90)
    for r in results:
        print(
            f"{r['icon']} {r['siteid']:<8} {r['score']:>5}/100  {r['category']:<15} "
            f"{r['subjects']:>5} {r['critical_queries']:>5} {r['sae_pending']:>5} "
            f"{r['stale_queries']:>6}  {r['action']}"
        )

    print("\n── Risk Factors Detail ─────────────────────────────────────")
    for r in results:
        if r["factors"]:
            print(f"\n  {r['icon']} {r['siteid']} (Score: {r['score']})")
            for f in r["factors"]:
                print(f"    [{f['level']:<8}] {f['factor']:<35} +{f['points']} pts")

    # ── Save JSON report ──────────────────────────────────────
    report = {
        "generated_at": datetime.now().isoformat(),
        "site_count": len(results),
        "high_risk": sum(1 for r in results if r["category"] == "HIGH RISK"),
        "medium_risk": sum(1 for r in results if r["category"] == "MEDIUM RISK"),
        "low_risk": sum(1 for r in results if r["category"] == "LOW RISK"),
        "sites": results,
    }
    json_path = os.path.join(REPORT_DIR, "risk_assessment.json")
    with open(json_path, "w") as f:
        json.dump(report, f, indent=2)
    print(f"\n[RISK] Report saved → {json_path}")

    # ── Try to save Excel if openpyxl available ───────────────
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Risk Assessment"

        # Header
        headers = [
            "Site ID",
            "Risk Score",
            "Category",
            "Action Required",
            "Subjects",
            "Critical Queries",
            "SAEs Pending",
            "Stale Queries",
            "Unans. Rate %",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0D2B55")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        fill_high = PatternFill("solid", fgColor="FFCCCC")
        fill_medium = PatternFill("solid", fgColor="FFF3CC")
        fill_low = PatternFill("solid", fgColor="CCFFCC")

        for row_n, r in enumerate(results, 2):
            fill = (
                fill_high
                if r["score"] >= 60
                else (fill_medium if r["score"] >= 30 else fill_low)
            )
            vals = [
                r["siteid"],
                r["score"],
                r["category"],
                r["action"],
                r["subjects"],
                r["critical_queries"],
                r["sae_pending"],
                r["stale_queries"],
                r["unanswered_rate"],
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row_n, column=col, value=v)
                cell.fill = fill

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 40

        xlsx_path = os.path.join(REPORT_DIR, "risk_assessment.xlsx")
        wb.save(xlsx_path)
        print(f"[RISK] Excel saved  → {xlsx_path}")
    except ImportError:
        print("[RISK] openpyxl not installed — skipping Excel export")

    print("\n" + "=" * 65)
    return results


if __name__ == "__main__":
    run_risk_assessment()
